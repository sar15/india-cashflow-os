from datetime import date

from cashflow_os.utils.dates import today_ist
from io import BytesIO, StringIO
from typing import Dict, List, Optional

import csv
from openpyxl import load_workbook

from cashflow_os.domain.models import (
    BankBalanceSnapshot,
    CanonicalCashEvent,
    Counterparty,
    EntityType,
    EventStatus,
    EventType,
    ImportBatch,
    ImportIssue,
    ParsedImportBundle,
    RecurringObligation,
    RelationshipType,
    Severity,
    SourceType,
)
from cashflow_os.utils.dates import parse_date_value
from cashflow_os.utils.money import parse_indian_number, to_minor_units


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _validated_date(
    value: object,
    *,
    issues: List[ImportIssue],
    row_number: Optional[int],
    field_name: str,
) -> Optional[date]:
    parsed_value = parse_date_value(value)
    if value not in (None, "") and parsed_value is None:
        issues.append(
            ImportIssue(
                code="invalid_date",
                severity=Severity.WARNING,
                message="Field '{field}' contains an invalid date value.".format(field=field_name),
                field_name=field_name,
                row_number=row_number,
            )
        )
    return parsed_value


def parse_manual_workbook(org_id: str, filename: str, file_bytes: bytes) -> ParsedImportBundle:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    batch = ImportBatch(org_id=org_id, source_type=SourceType.MANUAL, filename=filename)
    counterparties: Dict[str, Counterparty] = {}
    events: List[CanonicalCashEvent] = []
    obligations: List[RecurringObligation] = []
    issues: List[ImportIssue] = []
    bank_balance: Optional[BankBalanceSnapshot] = None

    if "Setup" in workbook.sheetnames:
        setup_sheet = workbook["Setup"]
        values = {
            str(row[0].value).strip().lower(): row[1].value
            for row in setup_sheet.iter_rows(min_row=2)
            if row[0].value is not None
        }
        opening_cash = values.get("opening_cash_inr") or values.get("opening_cash")
        if opening_cash is not None:
            bank_balance = BankBalanceSnapshot(
                org_id=org_id,
                as_of_date=_validated_date(
                    values.get("as_of_date"),
                    issues=issues,
                    row_number=None,
                    field_name="as_of_date",
                )
                or (workbook.properties.modified.date() if workbook.properties.modified else today_ist()),
                balance_minor_units=to_minor_units(parse_indian_number(opening_cash)),
            )

    if bank_balance is None:
        bank_balance = BankBalanceSnapshot(org_id=org_id, as_of_date=today_ist(), balance_minor_units=0)

    if "Cash Events" not in workbook.sheetnames:
        issues.append(
            ImportIssue(
                code="missing_sheet",
                severity=Severity.WARNING,
                message="Manual workbook is missing the 'Cash Events' sheet.",
            )
        )
    else:
        event_sheet = workbook["Cash Events"]
        headers = [_normalize_header(cell.value) for cell in next(event_sheet.iter_rows(min_row=1, max_row=1))[0:12]]
        for row_index, row in enumerate(event_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            row_data = {headers[index]: row[index] for index in range(min(len(headers), len(row)))}
            counterparty_name = str(row_data.get("counterparty") or row_data.get("entity_name") or "").strip() or None
            relationship = RelationshipType.CUSTOMER if str(row_data.get("event_type", "")).lower() == "inflow" else RelationshipType.VENDOR
            counterparty = None
            if counterparty_name:
                counterparty = counterparties.get(counterparty_name)
                if counterparty is None:
                    counterparty = Counterparty(
                        entity_name=counterparty_name,
                        relationship_type=relationship,
                        is_msme_registered=str(row_data.get("is_msme", "")).lower() in ("true", "yes", "1"),
                    )
                    counterparties[counterparty_name] = counterparty

            try:
                event_type = EventType(str(row_data.get("event_type") or "outflow").lower())
            except ValueError:
                issues.append(
                    ImportIssue(
                        code="invalid_event_type",
                        severity=Severity.WARNING,
                        message="Row {row} has an invalid event type.".format(row=row_index),
                        row_number=row_index,
                    )
                )
                continue

            entity_type_value = str(row_data.get("entity_type") or "manual").lower()
            entity_type = EntityType(entity_type_value) if entity_type_value in EntityType._value2member_map_ else EntityType.MANUAL

            amount = to_minor_units(parse_indian_number(row_data.get("gross_amount_inr") or row_data.get("amount_inr") or 0))
            tax_amount = to_minor_units(parse_indian_number(row_data.get("tax_amount_inr") or 0))
            tds_amount = to_minor_units(parse_indian_number(row_data.get("tds_amount_inr") or 0))
            net_amount = amount - tds_amount
            if net_amount < 0:
                net_amount = amount

            events.append(
                CanonicalCashEvent(
                    org_id=org_id,
                    source_id="manual.template",
                    import_batch_id=batch.import_batch_id,
                    event_type=event_type,
                    entity_type=entity_type,
                    counterparty_id=counterparty.counterparty_id if counterparty else None,
                    counterparty_name=counterparty_name,
                    document_number=str(row_data.get("document_number") or "MANUAL-{row}".format(row=row_index)),
                    document_date=_validated_date(row_data.get("document_date"), issues=issues, row_number=row_index, field_name="document_date"),
                    due_date=_validated_date(row_data.get("due_date"), issues=issues, row_number=row_index, field_name="due_date"),
                    expected_cash_date=_validated_date(
                        row_data.get("expected_cash_date"),
                        issues=issues,
                        row_number=row_index,
                        field_name="expected_cash_date",
                    ),
                    gross_minor_units=amount,
                    tax_minor_units=tax_amount,
                    tds_minor_units=tds_amount,
                    net_minor_units=net_amount,
                    status=EventStatus(str(row_data.get("status") or "open").lower())
                    if str(row_data.get("status") or "open").lower() in EventStatus._value2member_map_
                    else EventStatus.OPEN,
                    source_confidence=1.0,
                    mapping_confidence=0.95,
                    notes=str(row_data.get("notes") or ""),
                )
            )

    if "Obligations" in workbook.sheetnames:
        obligation_sheet = workbook["Obligations"]
        obligation_headers = [_normalize_header(cell.value) for cell in next(obligation_sheet.iter_rows(min_row=1, max_row=1))]
        for row in obligation_sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            row_data = {obligation_headers[index]: row[index] for index in range(min(len(obligation_headers), len(row)))}
            entity_type_value = str(row_data.get("obligation_type") or "other").lower()
            obligation_type = EntityType(entity_type_value) if entity_type_value in EntityType._value2member_map_ else EntityType.OTHER
            obligations.append(
                RecurringObligation(
                    org_id=org_id,
                    name=str(row_data.get("name") or obligation_type.value.title()),
                    obligation_type=obligation_type,
                    frequency=str(row_data.get("frequency") or "monthly").lower(),
                    amount_minor_units=to_minor_units(parse_indian_number(row_data.get("amount_inr") or 0)),
                    due_day=int(row_data.get("due_day") or 0) or None,
                    start_date=_validated_date(row_data.get("start_date"), issues=issues, row_number=None, field_name="start_date") or today_ist(),
                )
            )

    batch.event_count = len(events)
    batch.counterparty_count = len(counterparties)
    batch.obligation_count = len(obligations)
    batch.unresolved_issues = issues
    return ParsedImportBundle(
        import_batch=batch,
        bank_balance=bank_balance,
        counterparties=list(counterparties.values()),
        events=events,
        obligations=obligations,
    )


import csv

def parse_manual_csv(org_id: str, filename: str, file_text: str) -> ParsedImportBundle:
    reader = csv.reader(StringIO(file_text))
    workbook_bytes = BytesIO()
    
    from openpyxl import Workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cash Events"
    
    for row in reader:
        sheet.append(row)
        
    workbook.save(workbook_bytes)
    return parse_manual_workbook(org_id=org_id, filename=filename, file_bytes=workbook_bytes.getvalue())
