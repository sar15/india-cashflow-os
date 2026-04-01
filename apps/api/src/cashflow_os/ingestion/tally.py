from datetime import date
from io import BytesIO, StringIO
from typing import Dict, List, Optional
from xml.etree import ElementTree
import csv

from cashflow_os.utils.dates import today_ist, parse_date_value
from cashflow_os.utils.money import parse_indian_number, to_minor_units
from cashflow_os.domain.models import (
    BankBalanceSnapshot,
    CanonicalCashEvent,
    Counterparty,
    EntityType,
    EventType,
    ImportBatch,
    ImportIssue,
    ParsedImportBundle,
    RelationshipType,
    Severity,
    SourceType,
)

PARTY_HEADERS = {"party_name", "party", "ledger", "entity_name", "account_name", "name", "particulars"}
DOC_HEADERS = {"bill_no", "invoice_no", "document_number", "voucher_no", "ref_no"}
DATE_HEADERS = {"bill_date", "invoice_date", "document_date", "date"}
DUE_HEADERS = {"due_date", "expected_date"}
AMOUNT_HEADERS = {"amount", "balance", "pending_amount", "closing_balance", "outstanding_amount"}
MSME_HEADERS = {"is_msme", "msme", "msme_registered"}


def _normalize_key(key: str) -> str:
    return str(key or "").strip().lower().replace(" ", "_")


def _first_matching_row(row: Dict[str, str], candidates: set) -> Optional[str]:
    for candidate in candidates:
        if candidate in row and row[candidate] not in (None, ""):
            return row[candidate]
    return None


def _issue_invalid_date(issues: List[ImportIssue], row_number: int, field_name: str) -> None:
    issues.append(
        ImportIssue(
            code="invalid_date",
            severity=Severity.WARNING,
            message="Row {row} has an invalid {field} value.".format(row=row_number, field=field_name),
            row_number=row_number,
            field_name=field_name,
        )
    )


def _parse_xml_rows(file_bytes: bytes) -> List[Dict[str, str]]:
    root = ElementTree.fromstring(file_bytes)
    rows: List[Dict[str, str]] = []
    candidate_tags = ("row", "voucher", "ledgerentry", "ledger", "bill")

    for element in root.iter():
        if element.tag.lower() not in candidate_tags:
            continue
        row_data: Dict[str, str] = {}
        for child in element:
            text = (child.text or "").strip()
            if text:
                row_data[child.tag.lower()] = text
        if row_data:
            rows.append(row_data)

    if not rows:
        raise ValueError("No tabular rows found in the Tally XML export.")
    return rows


def _parse_frame(org_id: str, filename: str, frame: List[Dict[str, str]], source_hint: str) -> ParsedImportBundle:
    batch = ImportBatch(org_id=org_id, source_type=SourceType.TALLY, filename=filename)
    counterparties: Dict[str, Counterparty] = {}
    events: List[CanonicalCashEvent] = []
    issues: List[ImportIssue] = []
    event_type = EventType.INFLOW if source_hint == "receivables" else EventType.OUTFLOW
    relationship = RelationshipType.CUSTOMER if source_hint == "receivables" else RelationshipType.VENDOR
    entity_type = EntityType.INVOICE if source_hint == "receivables" else EntityType.BILL

    for row_index, row in enumerate(frame, start=2):
        party_name = _first_matching_row(row, PARTY_HEADERS)
        amount = _first_matching_row(row, AMOUNT_HEADERS)
        if not party_name or not amount:
            continue

        party_name_text = str(party_name).strip()
        counterparty = counterparties.get(party_name_text)
        if counterparty is None:
            msme_val = str(_first_matching_row(row, MSME_HEADERS) or "").lower()
            counterparty = Counterparty(
                entity_name=party_name_text,
                relationship_type=relationship,
                is_msme_registered=msme_val in ("true", "yes", "1"),
            )
            counterparties[party_name_text] = counterparty

        amount_value = to_minor_units(parse_indian_number(amount))
        document_number = str(_first_matching_row(row, DOC_HEADERS) or "{hint}-{row}".format(hint=source_hint, row=row_index))
        raw_document_date = _first_matching_row(row, DATE_HEADERS)
        raw_due_date = _first_matching_row(row, DUE_HEADERS)
        document_date = parse_date_value(raw_document_date)
        due_date = parse_date_value(raw_due_date)

        if raw_document_date not in (None, "") and document_date is None:
            _issue_invalid_date(issues, row_index, "document_date")
        if raw_due_date not in (None, "") and due_date is None:
            _issue_invalid_date(issues, row_index, "due_date")
        if due_date is None and document_date is None:
            issues.append(
                ImportIssue(
                    code="missing_schedule",
                    severity=Severity.WARNING,
                    message="Row {row} is missing both document date and due date.".format(row=row_index),
                    row_number=row_index,
                )
            )

        events.append(
            CanonicalCashEvent(
                org_id=org_id,
                source_id="tally.upload",
                import_batch_id=batch.import_batch_id,
                event_type=event_type,
                entity_type=entity_type,
                counterparty_id=counterparty.counterparty_id,
                counterparty_name=counterparty.entity_name,
                document_number=document_number,
                document_date=document_date,
                due_date=due_date,
                gross_minor_units=amount_value,
                net_minor_units=amount_value,
                source_confidence=0.9,
                mapping_confidence=0.85,
                notes="Parsed from Tally export",
            )
        )

    batch.event_count = len(events)
    batch.counterparty_count = len(counterparties)
    batch.unresolved_issues = issues
    return ParsedImportBundle(
        import_batch=batch,
        bank_balance=BankBalanceSnapshot(org_id=org_id, as_of_date=today_ist(), balance_minor_units=0),
        counterparties=list(counterparties.values()),
        events=events,
    )


def parse_tally_file(org_id: str, filename: str, file_bytes: bytes, source_hint: str) -> ParsedImportBundle:
    frame: List[Dict[str, str]] = []
    
    if filename.lower().endswith(".csv"):
        reader = csv.DictReader(StringIO(file_bytes.decode("utf-8")))
        if reader.fieldnames:
            normalized_fields = [_normalize_key(f) for f in reader.fieldnames]
            reader.fieldnames = normalized_fields
            frame = [dict(row) for row in reader]
    elif filename.lower().endswith(".xml") or file_bytes.lstrip().startswith(b"<"):
        frame = _parse_xml_rows(file_bytes)
    else:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
        sheet = wb.active
        headers = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                headers = [_normalize_key(h) for h in row]
            else:
                row_data = {headers[j]: str(val) if val is not None else "" for j, val in enumerate(row) if j < len(headers)}
                frame.append(row_data)

    return _parse_frame(org_id=org_id, filename=filename, frame=frame, source_hint=source_hint)
